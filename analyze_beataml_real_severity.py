from pathlib import Path
import warnings
import numpy as np
import pandas as pd
from scipy.stats import spearmanr, mannwhitneyu, pearsonr
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

PRED = Path("results/bulk_BeatAML/predictions_60pct.csv")
XLSX = Path("data/beataml_drugs/beataml_clinical.xlsx")
AUC  = Path("data/beataml_drugs/processed/beataml_key_drugs_auc.csv")

OUT = Path("results/markus_validation/real_clinical")
OUT.mkdir(parents=True, exist_ok=True)


def read_sheet(name):
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[name]

    rows = ws.iter_rows(values_only=True)
    first = next(rows)

    header = [
        str(x).strip() if x is not None else f"unnamed_{i}"
        for i, x in enumerate(first)
    ]

    df = pd.DataFrame(list(rows), columns=header)
    wb.close()

    return df


def num(x):
    return pd.to_numeric(x, errors="coerce")


def binary(v):
    if pd.isna(v):
        return np.nan

    if isinstance(v, (bool, np.bool_)):
        return int(v)

    s = str(v).strip().lower()

    positive = {
        "y", "yes", "true", "1",
        "positive", "pos",
        "mutated", "mutation",
        "detected", "present"
    }

    negative = {
        "n", "no", "false", "0",
        "negative", "neg",
        "wildtype", "wild-type", "wt",
        "not detected", "absent", "none"
    }

    if s in positive:
        return 1

    if s in negative:
        return 0

    if any(k in s for k in ["positive", "mutat", "detected", "present"]):
        return 1

    if any(k in s for k in ["negative", "wild", "not detected", "absent"]):
        return 0

    return np.nan


def bh(p):
    """Benjamini-Hochberg FDR."""
    p = np.asarray(p, dtype=float)

    order = np.argsort(p)
    ranked = p[order]

    q = ranked * len(p) / np.arange(1, len(p)+1)
    q = np.minimum.accumulate(q[::-1])[::-1]
    q = np.clip(q, 0, 1)

    out = np.empty_like(q)
    out[order] = q

    return out


def continuous_test(df, col, label, level):

    z = df[["persister_probability", col]].copy()
    z[col] = num(z[col])
    z = z.dropna()

    if len(z) < 20:
        return None

    rho, p = spearmanr(
        z["persister_probability"],
        z[col]
    )

    return {
        "level": level,
        "variable": label,
        "test": "Spearman",
        "n": len(z),
        "effect": rho,
        "p": p
    }


def binary_test(df, col, label, level):

    z = df[["persister_probability", col]].copy()
    z["_binary"] = z[col].map(binary)

    z = z.dropna(
        subset=["persister_probability", "_binary"]
    )

    if z["_binary"].nunique() != 2:
        return None

    negative = z.loc[
        z["_binary"] == 0,
        "persister_probability"
    ]

    positive = z.loc[
        z["_binary"] == 1,
        "persister_probability"
    ]

    if min(len(negative), len(positive)) < 5:
        return None

    U, p = mannwhitneyu(
        negative,
        positive,
        alternative="two-sided"
    )

    return {
        "level": level,
        "variable": label,
        "test": "Mann-Whitney",
        "n": len(z),
        "n0": len(negative),
        "n1": len(positive),
        "median0": negative.median(),
        "median1": positive.median(),
        "effect": positive.median() - negative.median(),
        "p": p
    }


# ============================================================
# VERIFIED PERSISTER SCORES
# ============================================================

pred = pd.read_csv(PRED)

pred = pred[
    pred["sample_id"].astype(str).str.lower()
    != "entrez_gene_id"
].copy()

# BeatAML sample:
# aml_ohsu_2018_12-00023 -> 12-00023
pred["labId"] = (
    pred["sample_id"]
    .astype(str)
    .str.extract(r"(\d{2}-\d{5})$")
)

pred = pred.dropna(
    subset=["labId", "persister_probability"]
)


# ============================================================
# REAL BEATAML CLINICAL SUMMARY
# ============================================================

print("Loading real BeatAML clinical summary...")

clin = read_sheet("Tabe S5-Clinical Summary")

clin["LabId"] = (
    clin["LabId"]
    .astype(str)
    .str.strip()
)

m = pred.merge(
    clin,
    left_on="labId",
    right_on="LabId",
    how="left",
    validate="one_to_one",
    suffixes=("", "_clinical")
)


print("\n" + "="*80)
print("REAL BEATAML MATCH")
print("="*80)

print("Prediction samples:", len(pred))
print(
    "Matched to clinical S5:",
    m["PatientId"].notna().sum()
)
print(
    "Unique matched patients:",
    m["PatientId"].nunique()
)

unmatched = m.loc[
    m["PatientId"].isna(),
    "sample_id"
].tolist()

print("Unmatched:", len(unmatched))

if unmatched:
    print(unmatched[:20])


# ============================================================
# SHOW AVAILABLE RELEVANT VARIABLES
# ============================================================

print("\n" + "="*80)
print("AVAILABLE CLINICAL VARIABLES")
print("="*80)

for c in m.columns:

    low = c.lower()

    if any(k in low for k in [
        "age",
        "blast",
        "eln",
        "relapse",
        "denovo",
        "transform",
        "specimen",
        "stage",
        "wbc",
        "white",
        "flt3",
        "npm1",
        "asxl1",
        "wes",
        "exome"
    ]):
        print(repr(c))


m.to_csv(
    OUT / "beataml_score_real_clinical_matched.csv",
    index=False
)


# ============================================================
# ONE SPECIMEN PER PATIENT SENSITIVITY SET
# ============================================================

if "timeOfSampleCollectionRelativeToInclusion" in m.columns:

    m["_collection_time"] = num(
        m["timeOfSampleCollectionRelativeToInclusion"]
    )

else:
    m["_collection_time"] = np.nan


m["_sort_time"] = (
    m["_collection_time"]
    .fillna(np.inf)
)


one = (
    m.dropna(subset=["PatientId"])
     .sort_values([
         "PatientId",
         "_sort_time",
         "labId"
     ])
     .drop_duplicates(
         "PatientId",
         keep="first"
     )
     .copy()
)


print(
    "\nEarliest matched specimen per patient:",
    len(one)
)


# ============================================================
# DISEASE-SEVERITY / CLINICAL ASSOCIATIONS
# ============================================================

results = []

continuous_variables = [

    (
        "ageAtDiagnosis",
        "Age at diagnosis"
    ),

    (
        "ageAtSpecimenAcquisition",
        "Age at specimen acquisition"
    ),

    (
        "%.Blasts.in.BM",
        "Bone-marrow blast %"
    ),

    (
        "%.Blasts.in.PB",
        "Peripheral-blood blast %"
    ),
]


binary_variables = [

    (
        "isRelapse",
        "Relapse"
    ),

    (
        "isDenovo",
        "De novo disease"
    ),

    (
        "isTransformed",
        "Transformed disease"
    ),

    (
        "FLT3-ITD",
        "FLT3-ITD"
    ),

    (
        "NPM1",
        "NPM1"
    ),

    (
        "ASXL1",
        "ASXL1"
    ),
]


for level, df in [

    ("specimen", m),

    (
        "one_specimen_per_patient",
        one
    )
]:

    for col, label in continuous_variables:

        if col in df.columns:

            r = continuous_test(
                df,
                col,
                label,
                level
            )

            if r:
                results.append(r)


    for col, label in binary_variables:

        if col in df.columns:

            r = binary_test(
                df,
                col,
                label,
                level
            )

            if r:
                results.append(r)


res = pd.DataFrame(results)

if len(res):

    res["q_BH"] = bh(res["p"].values)

    res = res.sort_values([
        "level",
        "p"
    ])

    print("\n" + "="*80)
    print("REAL CLINICAL ASSOCIATIONS")
    print("="*80)

    print(
        res.to_string(index=False)
    )

    res.to_csv(
        OUT / "real_clinical_associations.csv",
        index=False
    )


# ============================================================
# ELN2017 — SAME ORIGINAL DATASET
# ============================================================

if "ELN2017" in m.columns:

    e = m.dropna(
        subset=["ELN2017"]
    )

    print("\n" + "="*80)
    print("ELN2017 SCORE DISTRIBUTION")
    print("="*80)

    print(
        e.groupby("ELN2017")[
            "persister_probability"
        ]
        .agg([
            "count",
            "median",
            "mean"
        ])
        .to_string()
    )


# ============================================================
# SHOW EXACT ENCODING OF CLINICAL MUTATION CALLS
# ============================================================

print("\n" + "="*80)
print("CLINICAL MUTATION CALL ENCODINGS")
print("="*80)

for c in [
    "FLT3-ITD",
    "NPM1",
    "ASXL1"
]:

    if c in m.columns:

        print("\n", c)

        print(
            m[c]
            .value_counts(dropna=False)
            .head(20)
            .to_string()
        )


# ============================================================
# S7 REAL VARIANT DATA
# ============================================================

print("\nLoading S7 variant calls...")

variants = read_sheet(
    "Table S7-Variants for Analysis"
)

variants["labId"] = (
    variants["labId"]
    .astype(str)
    .str.strip()
)


genes = [
    "TP53",
    "RUNX1",
    "ASXL1",
    "FLT3",
    "NPM1",
    "DNMT3A",
    "IDH1",
    "IDH2",
    "TET2",
    "NRAS",
    "KRAS",
    "CEBPA"
]


variants["symbol"] = (
    variants["symbol"]
    .astype(str)
    .str.upper()
)


v = variants[
    variants["symbol"].isin(genes)
][
    ["labId", "symbol"]
].drop_duplicates()


vp = (
    m[
        [
            "sample_id",
            "labId",
            "PatientId",
            "persister_probability"
        ]
    ]
    .merge(
        v.assign(mutation_present=1),
        on="labId",
        how="left"
    )
)


print("\n" + "="*80)
print("S7 MUTATION-POSITIVE COUNTS")
print("="*80)

print(
    "Important: these are descriptive counts only."
)
print(
    "We do NOT yet treat absence from S7 as wild-type "
    "until WES eligibility is verified."
)

mutation_counts = (
    vp.dropna(subset=["mutation_present"])
      .groupby("symbol")["labId"]
      .nunique()
      .sort_values(ascending=False)
)

print(mutation_counts.to_string())


# ============================================================
# PARTIAL SPEARMAN
# Drug resistance adjusted for ELN + age
# ============================================================

def partial_spearman(x, y, covariates):

    d = pd.DataFrame({
        "x": x,
        "y": y
    }).join(covariates)

    d = d.dropna()

    if len(d) < 20:
        return np.nan, np.nan, len(d)

    # Spearman = Pearson correlation of ranks
    rx = d["x"].rank().to_numpy(float)
    ry = d["y"].rank().to_numpy(float)

    cov = d.drop(
        columns=["x", "y"]
    )

    cov = pd.get_dummies(
        cov,
        drop_first=True,
        dtype=float
    )

    X = np.column_stack([
        np.ones(len(d)),
        cov.to_numpy(float)
    ])

    bx = np.linalg.lstsq(
        X,
        rx,
        rcond=None
    )[0]

    by = np.linalg.lstsq(
        X,
        ry,
        rcond=None
    )[0]

    ex = rx - X @ bx
    ey = ry - X @ by

    r, p = pearsonr(ex, ey)

    return r, p, len(d)


if AUC.exists() and "ELN2017" in m.columns:

    auc = pd.read_csv(AUC)

    auc = auc[
        auc["sample_id"]
        .astype(str)
        .str.lower()
        != "entrez_gene_id"
    ].copy()


    keep = [
        "sample_id",
        "persister_probability",
        "ELN2017"
    ]

    if "ageAtDiagnosis" in m.columns:
        keep.append(
            "ageAtDiagnosis"
        )


    da = m[keep].merge(
        auc,
        on="sample_id",
        how="inner"
    )


    drug_columns = [
        c for c in auc.columns
        if c != "sample_id"
    ]


    adjusted = []


    for drug in drug_columns:

        if drug not in da.columns:
            continue

        cov = pd.DataFrame({
            "ELN2017":
                da["ELN2017"]
                .astype("category")
        })


        if "ageAtDiagnosis" in da.columns:

            cov["ageAtDiagnosis"] = num(
                da["ageAtDiagnosis"]
            )


        r, p, n = partial_spearman(
            da["persister_probability"],
            num(da[drug]),
            cov
        )


        if np.isfinite(r):

            adjusted.append({
                "drug": drug,
                "n": n,
                "partial_spearman": r,
                "p": p
            })


    adjusted = pd.DataFrame(adjusted)


    if len(adjusted):

        adjusted["q_BH"] = bh(
            adjusted["p"].values
        )

        print("\n" + "="*80)
        print(
            "DRUG AUC — PARTIAL SPEARMAN "
            "ADJUSTED FOR ELN2017 + AGE"
        )
        print("="*80)

        print(
            adjusted.to_string(
                index=False
            )
        )


        adjusted.to_csv(
            OUT /
            "drug_auc_partial_spearman_eln_age.csv",
            index=False
        )


print("\n" + "="*80)
print("DONE")
print("="*80)
print("Outputs:", OUT)
